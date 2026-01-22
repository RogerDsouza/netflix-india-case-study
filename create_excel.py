from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

print("Creating Excel financial model...")

wb = Workbook()

# Sheet 1: Pricing Optimization
ws1 = wb.active
ws1.title = "Pricing Optimization"

# Header
ws1['A1'] = "NETFLIX INDIA - PRICING OPTIMIZATION MODEL"
ws1['A1'].font = Font(size=14, bold=True)
ws1.merge_cells('A1:H1')

# Assumptions
ws1['A3'] = "KEY ASSUMPTIONS"
ws1['A3'].font = Font(size=12, bold=True, color='0000FF')

ws1['A4'] = "Current Price (Mobile Plan)"
ws1['B4'] = 2.85
ws1['B4'].font = Font(color='0000FF')

ws1['A5'] = "Current Subscribers (millions)"
ws1['B5'] = 2.0
ws1['B5'].font = Font(color='0000FF')

ws1['A6'] = "Price Elasticity"
ws1['B6'] = -1.5
ws1['B6'].font = Font(color='0000FF')

ws1['A7'] = "Variable Cost per User"
ws1['B7'] = 0.50
ws1['B7'].font = Font(color='0000FF')

ws1['A8'] = "Fixed Costs (millions/month)"
ws1['B8'] = 1.0
ws1['B8'].font = Font(color='0000FF')

# Price optimization table
ws1['A11'] = "PRICE OPTIMIZATION ANALYSIS"
ws1['A11'].font = Font(size=12, bold=True)

headers = ['Price', 'Price Change %', 'Demand Change %', 'Est. Subscribers (M)', 
           'Revenue (M)', 'Variable Costs (M)', 'Contribution Margin (M)', 'Profit (M)']
for col, header in enumerate(headers, start=1):
    cell = ws1.cell(row=12, column=col)
    cell.value = header
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color='D5E8F0', end_color='D5E8F0', fill_type='solid')

# Price points
prices = [1.99, 2.50, 2.85, 3.50, 4.00, 4.50, 5.00]
for idx, price in enumerate(prices, start=13):
    ws1.cell(row=idx, column=1, value=price)
    ws1.cell(row=idx, column=2, value=f'=(A{idx}-$B$4)/$B$4')
    ws1.cell(row=idx, column=3, value=f'=B{idx}*$B$6')
    ws1.cell(row=idx, column=4, value=f'=$B$5*(1+C{idx})')
    ws1.cell(row=idx, column=5, value=f'=A{idx}*D{idx}')
    ws1.cell(row=idx, column=6, value=f'=$B$7*D{idx}')
    ws1.cell(row=idx, column=7, value=f'=E{idx}-F{idx}')
    ws1.cell(row=idx, column=8, value=f'=G{idx}-$B$8')

ws1['A21'] = "OPTIMAL PRICE:"
ws1['A21'].font = Font(size=11, bold=True, color='FF0000')
ws1['B21'] = '=INDEX(A13:A19,MATCH(MAX(H13:H19),H13:H19,0))'
ws1['B21'].font = Font(size=11, bold=True, color='FF0000')

ws1['A22'] = "MAX PROFIT:"
ws1['B22'] = '=MAX(H13:H19)'
ws1['B22'].font = Font(bold=True, color='FF0000')

# Sheet 2: CLV Analysis with Limits
ws2 = wb.create_sheet("CLV Analysis")

ws2['A1'] = "CUSTOMER LIFETIME VALUE - CALCULUS APPLICATION"
ws2['A1'].font = Font(size=14, bold=True)
ws2.merge_cells('A1:F1')

ws2['A3'] = "ASSUMPTIONS"
ws2['A3'].font = Font(size=12, bold=True, color='0000FF')

ws2['A4'] = "Monthly ARPU ($)"
ws2['B4'] = 2.85
ws2['B4'].font = Font(color='0000FF')

ws2['A5'] = "Monthly Churn Rate (%)"
ws2['B5'] = 0.12
ws2['B5'].font = Font(color='0000FF')

ws2['A6'] = "Gross Margin (%)"
ws2['B6'] = 0.80
ws2['B6'].font = Font(color='0000FF')

ws2['A7'] = "Monthly Discount Rate (%)"
ws2['B7'] = 0.01
ws2['B7'].font = Font(color='0000FF')

ws2['A8'] = "CAC per Customer ($)"
ws2['B8'] = 5.00
ws2['B8'].font = Font(color='0000FF')

ws2['A10'] = "CLV CALCULATION (Using Limits)"
ws2['A10'].font = Font(size=11, bold=True)

ws2['A11'] = "Month"
ws2['B11'] = "Retention Rate"
ws2['C11'] = "Expected Customers"
ws2['D11'] = "Revenue"
ws2['E11'] = "Discounted Value"
ws2['F11'] = "Cumulative CLV"

for col in range(1, 7):
    ws2.cell(row=11, column=col).font = Font(bold=True)
    ws2.cell(row=11, column=col).fill = PatternFill(start_color='D5E8F0', end_color='D5E8F0', fill_type='solid')

for month in range(0, 25):
    row = 12 + month
    ws2.cell(row=row, column=1, value=month)
    ws2.cell(row=row, column=2, value=f'=(1-$B$5)^A{row}')
    ws2.cell(row=row, column=3, value=f'=B{row}*100')
    ws2.cell(row=row, column=4, value=f'=C{row}*$B$4*$B$6')
    ws2.cell(row=row, column=5, value=f'=D{row}/((1+$B$7)^A{row})')
    if month == 0:
        ws2.cell(row=row, column=6, value=f'=E{row}')
    else:
        ws2.cell(row=row, column=6, value=f'=F{row-1}+E{row}')

ws2['A39'] = "Theoretical CLV (Limit as n→∞):"
ws2['A39'].font = Font(bold=True, color='FF0000')
ws2['B39'] = '=$B$4*$B$6/($B$5+$B$7)'
ws2['B39'].font = Font(color='FF0000')

ws2['A40'] = "CLV from 24-month calculation:"
ws2['B40'] = '=F36'

ws2['A42'] = "Net CLV (CLV - CAC):"
ws2['A42'].font = Font(bold=True, color='0000FF')
ws2['B42'] = '=B39-$B$8'
ws2['B42'].font = Font(bold=True, color='0000FF')

ws2['A43'] = "LTV:CAC Ratio:"
ws2['B43'] = '=B39/$B$8'
ws2['B43'].font = Font(bold=True, color='0000FF')

# Sheet 3: Market Sizing
ws3 = wb.create_sheet("Market Sizing")

ws3['A1'] = "INDIA MARKET OPPORTUNITY - TAM ANALYSIS"
ws3['A1'].font = Font(size=14, bold=True)
ws3.merge_cells('A1:G1')

ws3['A3'] = "MARKET DATA"
ws3['A3'].font = Font(size=12, bold=True)

segments = [
    ('Premium Urban', 'Tier_1', 35, 10.0, 0.15),
    ('Mid Urban', 'Tier_1', 55, 5.0, 0.10),
    ('Value Urban', 'Tier_2', 80, 2.5, 0.08),
    ('Small City', 'Tier_3', 90, 1.5, 0.05),
    ('Rural Digital', 'Rural', 40, 0.8, 0.02)
]

headers = ['Segment', 'Tier', 'Viewers (M)', 'WTP ($)', 'Conversion %', 'Potential Subs (M)', 'Annual Revenue (M)']
for col, header in enumerate(headers, start=1):
    cell = ws3.cell(row=4, column=col)
    cell.value = header
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')

for idx, (segment, tier, viewers, wtp, conversion) in enumerate(segments, start=5):
    ws3.cell(row=idx, column=1, value=segment)
    ws3.cell(row=idx, column=2, value=tier)
    ws3.cell(row=idx, column=3, value=viewers)
    ws3.cell(row=idx, column=4, value=wtp)
    ws3.cell(row=idx, column=5, value=conversion)
    ws3.cell(row=idx, column=6, value=f'=C{idx}*E{idx}')
    ws3.cell(row=idx, column=7, value=f'=F{idx}*D{idx}*12')

ws3['A11'] = "TOTAL MARKET OPPORTUNITY"
ws3['A11'].font = Font(bold=True)
ws3['B11'] = '=SUM(F5:F9)'
ws3['B11'].font = Font(bold=True)
ws3['C11'] = 'million subscribers'

ws3['A12'] = "TOTAL REVENUE POTENTIAL"
ws3['B12'] = '=SUM(G5:G9)'
ws3['B12'].font = Font(bold=True, color='FF0000')
ws3['C12'] = 'million USD/year'

# Sheet 4: Churn Impact
ws4 = wb.create_sheet("Churn Impact")

ws4['A1'] = "CHURN RATE IMPACT ON CLV - SENSITIVITY ANALYSIS"
ws4['A1'].font = Font(size=14, bold=True)
ws4.merge_cells('A1:E1')

ws4['A3'] = "BASE ASSUMPTIONS"
ws4['A4'] = "ARPU"
ws4['B4'] = '=CLV!B4'
ws4['A5'] = "Gross Margin"
ws4['B5'] = '=CLV!B6'
ws4['A6'] = "Discount Rate"
ws4['B6'] = '=CLV!B7'

ws4['A8'] = "Churn Rate"
ws4['B8'] = "CLV ($)"
ws4['C8'] = "Change in CLV"
ws4['D8'] = "Marginal Impact (ΔCLV/Δchurn)"

for col in range(1, 5):
    ws4.cell(row=8, column=col).font = Font(bold=True)
    ws4.cell(row=8, column=col).fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')

churn_rates = [i/100 for i in range(5, 21)]
for idx, churn in enumerate(churn_rates, start=9):
    ws4.cell(row=idx, column=1, value=churn)
    ws4.cell(row=idx, column=2, value=f'=$B$4*$B$5/(A{idx}+$B$6)')
    if idx == 9:
        ws4.cell(row=idx, column=3, value=0)
        ws4.cell(row=idx, column=4, value=0)
    else:
        ws4.cell(row=idx, column=3, value=f'=B{idx}-B{idx-1}')
        ws4.cell(row=idx, column=4, value=f'=C{idx}/(A{idx}-A{idx-1})')

# Formatting
for ws in [ws1, ws2, ws3, ws4]:
    for col in range(1, 15):
        ws.column_dimensions[get_column_letter(col)].width = 18

# Number formatting
for row in range(13, 20):
    for col in [2, 3]:
        ws1.cell(row=row, column=col).number_format = '0.00%'
    for col in [4, 5, 6, 7, 8]:
        ws1.cell(row=row, column=col).number_format = '#,##0.00'

ws1['B21'].number_format = '$#,##0.00'
ws1['B22'].number_format = '$#,##0.00'

for col in [4, 5, 6]:
    for row in range(12, 37):
        ws2.cell(row=row, column=col).number_format = '$#,##0.00'

ws2['B39'].number_format = '$#,##0.00'
ws2['B40'].number_format = '$#,##0.00'
ws2['B42'].number_format = '$#,##0.00'
ws2['B43'].number_format = '0.00'

for row in range(5, 10):
    ws3.cell(row=row, column=3).number_format = '#,##0.0'
    ws3.cell(row=row, column=4).number_format = '$#,##0.00'
    ws3.cell(row=row, column=5).number_format = '0.0%'
    ws3.cell(row=row, column=6).number_format = '#,##0.00'
    ws3.cell(row=row, column=7).number_format = '$#,##0.00'

for row in range(9, 25):
    ws4.cell(row=row, column=1).number_format = '0.0%'
    ws4.cell(row=row, column=2).number_format = '$#,##0.00'
    ws4.cell(row=row, column=3).number_format = '$#,##0.00'
    ws4.cell(row=row, column=4).number_format = '$#,##0.00'

wb.save('analysis/netflix_india_financial_model.xlsx')
print("✅ Excel file created: analysis/netflix_india_financial_model.xlsx")